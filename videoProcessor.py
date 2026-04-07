"""
Avatar Training Data Pipeline
==============================
Full end-to-end pipeline for one video:
  1. Extract pose landmarks (MediaPipe) → temp CSV
  2. Extract audio → temp WAV
  3. Analyze audio (energy, pitch, mood) → temp XLSX
  4. Combine landmarks + audio features → final combined XLSX
  5. Clean up all temp files

SETUP — install dependencies:
    pip install opencv-python mediapipe moviepy librosa numpy pandas openpyxl

CONFIGURE:
    Set MODEL_PATH and OUTPUT_DIR below before running.

USAGE:
    python avatar_pipeline.py path/to/your_video.mp4
"""

import os
import sys
import csv
import argparse
import tempfile
from pathlib import Path

import cv2
import mediapipe as mp
import numpy as np
import pandas as pd
import librosa
from moviepy import VideoFileClip
from mediapipe.tasks.python import vision
from mediapipe.tasks.python import BaseOptions
from mediapipe.tasks.python.vision.pose_landmarker import PoseLandmarksConnections
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# CONFIGURE THESE TWO PATHS BEFORE RUNNING
# ─────────────────────────────────────────────
MODEL_PATH = "/Users/ryandepalma/Desktop/DCDS/MP/pose_landmarker_full.task"
OUTPUT_DIR = "./Output"
# ─────────────────────────────────────────────

LANDMARK_NAMES = [
    "NOSE", "LEFT_EYE_INNER", "LEFT_EYE", "LEFT_EYE_OUTER",
    "RIGHT_EYE_INNER", "RIGHT_EYE", "RIGHT_EYE_OUTER",
    "LEFT_EAR", "RIGHT_EAR", "MOUTH_LEFT", "MOUTH_RIGHT",
    "LEFT_SHOULDER", "RIGHT_SHOULDER", "LEFT_ELBOW", "RIGHT_ELBOW",
    "LEFT_WRIST", "RIGHT_WRIST", "LEFT_PINKY", "RIGHT_PINKY",
    "LEFT_INDEX", "RIGHT_INDEX", "LEFT_THUMB", "RIGHT_THUMB",
    "LEFT_HIP", "RIGHT_HIP", "LEFT_KNEE", "RIGHT_KNEE",
    "LEFT_ANKLE", "RIGHT_ANKLE", "LEFT_HEEL", "RIGHT_HEEL",
    "LEFT_FOOT_INDEX", "RIGHT_FOOT_INDEX"
]

KEEP_LANDMARKS = [
    'LEFT_SHOULDER', 'RIGHT_SHOULDER',
    'LEFT_ELBOW',    'RIGHT_ELBOW'
]

RENAME_MAP = {
    'LEFT_SHOULDER_x':  'L_Shoulder_x', 'LEFT_SHOULDER_y':  'L_Shoulder_y', 'LEFT_SHOULDER_z':  'L_Shoulder_z',
    'LEFT_ELBOW_x':     'L_Elbow_x',    'LEFT_ELBOW_y':     'L_Elbow_y',    'LEFT_ELBOW_z':     'L_Elbow_z',
    'RIGHT_SHOULDER_x': 'R_Shoulder_x', 'RIGHT_SHOULDER_y': 'R_Shoulder_y', 'RIGHT_SHOULDER_z': 'R_Shoulder_z',
    'RIGHT_ELBOW_x':    'R_Elbow_x',    'RIGHT_ELBOW_y':    'R_Elbow_y',    'RIGHT_ELBOW_z':    'R_Elbow_z',
}

HEADER_COLORS = {
    'timestamp_ms': 'FF4472C4',
    'Energy': 'FF70AD47', 'Pitch': 'FF70AD47', 'Mood': 'FF70AD47',
    'L_Shoulder_x': 'FFED7D31', 'L_Shoulder_y': 'FFED7D31', 'L_Shoulder_z': 'FFED7D31',
    'L_Elbow_x':    'FFED7D31', 'L_Elbow_y':    'FFED7D31', 'L_Elbow_z':    'FFED7D31',
    'R_Shoulder_x': 'FFFFC000', 'R_Shoulder_y': 'FFFFC000', 'R_Shoulder_z': 'FFFFC000',
    'R_Elbow_x':    'FFFFC000', 'R_Elbow_y':    'FFFFC000', 'R_Elbow_z':    'FFFFC000',
    'pose_detected': 'FFD9534F',  # red — quality flag
}


# ── Step 1: Extract pose landmarks ──────────────────────────────────────────

def extract_landmarks(video_path: Path, csv_path: Path):
    print("  [1/4] Extracting pose landmarks...")

    CONNECTIONS      = [(c.start, c.end) for c in PoseLandmarksConnections.POSE_LANDMARKS]
    HEAD_CONNECTIONS = [(0,7),(0,8),(7,11),(8,12),(9,10),(0,9),(0,10)]

    pose_detector = vision.PoseLandmarker.create_from_options(
        vision.PoseLandmarkerOptions(
            base_options=BaseOptions(model_asset_path=MODEL_PATH),
            running_mode=vision.RunningMode.VIDEO
        )
    )

    vid_capture = cv2.VideoCapture(str(video_path))
    if not vid_capture.isOpened():
        raise RuntimeError(f"Could not open video: {video_path}")

    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['timestamp_ms', 'pose_detected'] + [f"{n}_{a}" for n in LANDMARK_NAMES for a in ['x','y','z']])

        total_frames = 0
        detected_frames = 0

        while vid_capture.isOpened():
            frame_read, curr_frame = vid_capture.read()
            if not frame_read:
                break

            total_frames += 1
            timestamp_ms = int(vid_capture.get(cv2.CAP_PROP_POS_MSEC))
            mp_image = mp.Image(
                image_format=mp.ImageFormat.SRGB,
                data=cv2.cvtColor(curr_frame, cv2.COLOR_BGR2RGB)
            )
            results = pose_detector.detect_for_video(mp_image, timestamp_ms)

            if results.pose_landmarks:
                detected_frames += 1
                lm = results.pose_landmarks[0]
                row = [timestamp_ms, True]
                for l in lm:
                    row.extend([l.x, l.y, l.z] if l.visibility > 0.5 else [None, None, None])
            else:
                # Write the frame with pose_detected=False and all landmark columns blank
                row = [timestamp_ms, False] + [None] * (len(LANDMARK_NAMES) * 3)

            writer.writerow(row)

    vid_capture.release()
    pose_detector.close()
    missing = total_frames - detected_frames
    print(f"      {detected_frames}/{total_frames} frames with pose detected ({missing} missing).")


# ── Step 2: Extract audio ────────────────────────────────────────────────────

def extract_audio(video_path: Path, wav_path: Path):
    print("  [2/4] Extracting audio...")
    video = VideoFileClip(str(video_path))
    video.audio.write_audiofile(str(wav_path), logger=None)
    video.close()
    print(f"      Audio saved to temp file.")


# ── Step 3: Analyze audio features ──────────────────────────────────────────

def analyze_audio(wav_path: Path) -> pd.DataFrame:
    print("  [3/4] Analyzing audio features (energy, pitch, mood)...")

    y, sr = librosa.load(str(wav_path))
    samples_per_segment = sr  # 1 second

    results = []
    for i in range(0, len(y), samples_per_segment):
        segment = y[i:i + samples_per_segment]
        if len(segment) < sr:
            continue

        start_time = i / sr
        end_time   = (i + samples_per_segment) / sr
        energy     = float(np.mean(librosa.feature.rms(y=segment)))

        pitches, magnitudes = librosa.piptrack(y=segment, sr=sr)
        pitch_values = pitches[magnitudes > np.median(magnitudes)]
        avg_pitch = float(np.mean(pitch_values)) if len(pitch_values) > 0 else 0.0

        if energy > 0.02 and avg_pitch > 150:
            mood = "Engaged"
        elif energy < 0.01:
            mood = "Calm"
        else:
            mood = "Moderate"

        results.append({
            "Start_Time": start_time,
            "End_Time":   end_time,
            "Energy":     energy,
            "Pitch":      avg_pitch,
            "Mood":       mood
        })

    df = pd.DataFrame(results)
    df['start_ms'] = df['Start_Time'] * 1000
    df['end_ms']   = df['End_Time']   * 1000
    print(f"      {len(df)} audio segments analyzed.")
    return df


# ── Step 4: Combine landmarks + audio ───────────────────────────────────────

def combine(csv_path: Path, df_audio: pd.DataFrame, output_path: Path):
    print("  [4/4] Combining landmarks + audio features...")

    df_lm = pd.read_csv(csv_path)

    # Keep only shoulder and elbow columns plus pose_detected flag
    keep_cols = ['timestamp_ms', 'pose_detected'] + [
        f"{name}_{axis}"
        for name in KEEP_LANDMARKS
        for axis in ['x', 'y', 'z']
        if f"{name}_{axis}" in df_lm.columns
    ]
    df_lm = df_lm[keep_cols].copy()

    # Broadcast audio features onto each 41ms landmark frame
    def lookup(ts):
        match = df_audio[(df_audio['start_ms'] <= ts) & (df_audio['end_ms'] > ts)]
        if len(match):
            r = match.iloc[0]
            return r['Energy'], r['Pitch'], r['Mood']
        return None, None, None

    df_lm[['Energy', 'Pitch', 'Mood']] = df_lm['timestamp_ms'].apply(
        lambda t: pd.Series(lookup(t))
    )

    # Drop tail frames with no matching audio
    before = len(df_lm)
    df_lm = df_lm.dropna(subset=['Energy', 'Pitch', 'Mood'])
    dropped = before - len(df_lm)
    if dropped:
        print(f"      Dropped {dropped} tail frames with no audio match.")

    # Summary of pose detection coverage
    total = len(df_lm)
    detected = df_lm['pose_detected'].sum()
    print(f"      pose_detected: {detected}/{total} frames ({100*detected//total}% coverage)")

    # Reorder and rename columns — pose_detected comes last as a quality flag
    landmark_cols = [c for c in keep_cols if c not in ('timestamp_ms', 'pose_detected')]
    col_order = ['timestamp_ms', 'Energy', 'Pitch', 'Mood'] + landmark_cols + ['pose_detected']
    final = df_lm[col_order].rename(columns=RENAME_MAP)

    final.to_excel(output_path, index=False)
    style_output(output_path)
    print(f"      Combined file: {output_path}  ({len(final)} rows)")


# ── Styling ──────────────────────────────────────────────────────────────────

def style_output(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Combined"
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=HEADER_COLORS.get(cell.value, 'FFD9D9D9'))
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            max(len(str(cell.value or '')) for cell in col) + 2, 13)
    ws.freeze_panes = 'A2'
    wb.save(path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Full avatar training data pipeline for one video.')
    parser.add_argument('video', help='Path to the input .mp4 video file')
    args = parser.parse_args()

    video_path = Path(args.video)
    if not video_path.exists():
        print(f"ERROR: Video file not found: {video_path}")
        sys.exit(1)

    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    stem        = video_path.stem
    output_path = output_dir / f"{stem}_combined.xlsx"

    # Temp files — stored in system temp dir, cleaned up after
    tmp_dir  = Path(tempfile.mkdtemp())
    csv_path = tmp_dir / f"{stem}_landmarks.csv"
    wav_path = tmp_dir / f"{stem}.wav"

    print(f"\nProcessing: {video_path.name}")
    print("─" * 50)

    try:
        extract_landmarks(video_path, csv_path)
        extract_audio(video_path, wav_path)
        df_audio = analyze_audio(wav_path)
        combine(csv_path, df_audio, output_path)
    finally:
        # Always clean up temp files even if something fails
        for f in [csv_path, wav_path]:
            if f.exists():
                f.unlink()
        tmp_dir.rmdir()

    print("─" * 50)
    print(f"Done! Output saved to: {output_path.resolve()}\n")


if __name__ == '__main__':
    main()
